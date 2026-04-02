"""
GIKI Scheduler v2 - Course List Parser (Fixed)
================================================
Fixes:
  - Teacher deduplication: strip -1/-2 suffixes, normalize "New Engr." variants
  - Batch year from semester number: Sem2=Y1/2025, Sem4=Y2/2024, Sem6=Y3/2023, Sem8=Y4/2022
  - Section UID always = course_code + "-" + section_id
  - Handles all 8 GIKI Excel sheet formats
"""
from __future__ import annotations
import re, os
from dataclasses import dataclass, field
from typing import Optional


@dataclass
class CourseEntry:
    course_code:  str
    title:        str
    credit_hours: int
    instructor:   str
    section:      str
    faculty:      str
    semester:     int = 0
    batch_year:   int = 0   # 1=2025, 2=2024, 3=2023, 4=2022
    intake_year:  int = 0   # actual year: 2025/2024/2023/2022
    is_lab:       bool = False
    for_program:  str = ""
    num_students: int = 30
    raw_ch:       str = ""

    @property
    def uid(self) -> str:
        return f"{self.course_code}-{self.section}"


def _clean(val) -> str:
    return "" if val is None else str(val).strip()


def _parse_ch(ch_val) -> tuple[int, str]:
    raw = _clean(ch_val).upper().replace("CH", "").strip()
    if not raw:
        return 2, ""
    # L-T-P format: take first number (lectures)
    parts = re.split(r'[-/]', raw)
    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) >= 1:
        try:
            return max(1, int(float(parts[0]))), _clean(ch_val)
        except ValueError:
            pass
    try:
        return max(1, int(float(raw.split()[0]))), _clean(ch_val)
    except (ValueError, IndexError):
        return 2, _clean(ch_val)


def _sem_to_batch(sem) -> tuple[int, int]:
    """Convert semester number to (batch_year_num, intake_year)."""
    mapping = {2: (1, 2025), 4: (2, 2024), 6: (3, 2023), 8: (4, 2022)}
    try:
        s = int(str(sem).strip().replace("th","").replace("nd","").replace("rd","").replace("st",""))
        return mapping.get(s, (0, 0))
    except (ValueError, TypeError):
        return (0, 0)


def _infer_faculty(code: str) -> str:
    code = code.strip().upper()
    for prefixes, fac in [
        (("CS","CE","AI","CY","DS","SE","IF"), "FCSE"),
        (("EE",), "FEE"),
        (("ME",), "FME"),
        (("CH",), "FChE"),
        (("MM","CME","MTE"), "FMCE"),
        (("CV",), "FCvE"),
        (("MT","ES","PH"), "FBS"),
        (("HM","MS","AF","EM","SC"), "SMgS"),
    ]:
        for p in prefixes:
            if code.startswith(p):
                return fac
    return "FCSE"


def normalize_teacher_name(raw: str) -> tuple[str, str]:
    """
    Returns (canonical_name, teacher_id).
    Strips: -1 -2 -3 suffixes, (2-CH), (AI/CS), etc.
    Deduplicates: 'Dr. Memoon Sajid-1' and 'Dr. Memoon Sajid-2' -> same person.
    """
    name = _clean(raw)
    # Strip section suffixes like "-1", "-2", "-3" at end
    name = re.sub(r'\s*-\s*\d+\s*$', '', name)
    # Strip parenthetical notes: (2-CH), (AI/CS), (EE), etc.
    name = re.sub(r'\s*\([^)]*\)\s*$', '', name)
    # Normalize whitespace
    name = re.sub(r'\s+', ' ', name).strip()
    # Remove title prefixes for ID but keep in display name
    id_base = re.sub(r'^(Dr\.|Prof\.|Mr\.|Ms\.|Mrs\.|Engr\.|Miss\.)\s*', '', name, flags=re.I)
    id_base = re.sub(r'\s+', '-', id_base.strip()).upper()[:24]
    # Replace unusual chars
    id_base = re.sub(r'[^A-Z0-9\-]', '', id_base)
    return name, f"T-{id_base}"


def parse_excel(filepath: str) -> list[CourseEntry]:
    from openpyxl import load_workbook
    entries = []
    wb = load_workbook(filepath, read_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        entries.extend(_parse_sheet(rows, sheet_name))

    return entries


def _parse_sheet(rows: list, sheet_name: str) -> list[CourseEntry]:
    entries = []
    # Find header row
    header_idx = None
    col = {}
    for i, row in enumerate(rows):
        vals = [_clean(v).lower() for v in row]
        # Look for a row that has at least "code" and either "instructor" or "title"
        has_code = any("code" in v or "course#" in v for v in vals)
        has_instructor = any("instructor" in v or "instr" in v for v in vals)
        has_title = any("title" in v for v in vals)
        if has_code and (has_instructor or has_title):
            header_idx = i
            for j, v in enumerate(vals):
                if not v:
                    continue
                if ("course" in v and "code" in v) or v == "code" or "course#" in v:
                    col.setdefault("code", j)
                if "title" in v:
                    col.setdefault("title", j)
                if "credit" in v or (v == "ch" and "credit" not in col):
                    col.setdefault("ch", j)
                if "instructor" in v or "instr" in v:
                    col.setdefault("instructor", j)
                if v == "section" or v == "sec." or v == "sec":
                    col.setdefault("section", j)
                if "for" in v or "program" in v:
                    col.setdefault("for", j)
                if "sem" in v:
                    col.setdefault("sem", j)
                if "exp" in v or "student" in v or "strength" in v:
                    col.setdefault("students", j)
            break

    # Fallback: SMGS style (S.No, Code, Title, CH, Section, Instructor, For)
    if header_idx is None:
        for i, row in enumerate(rows):
            vals = [_clean(v) for v in row]
            if len(vals) >= 5 and any(re.match(r'^[A-Z]{2,3}\d{3}', _clean(v)) for v in vals):
                # Guess the layout
                for j, v in enumerate(vals):
                    v2 = _clean(row[j]).upper()
                    if re.match(r'^[A-Z]{2,3}\d{3}', v2) and "code" not in col:
                        col["code"] = j
                        if j+1 < len(vals): col.setdefault("title", j+1)
                        if j+2 < len(vals): col.setdefault("ch", j+2)
                        if j+3 < len(vals): col.setdefault("section", j+3)
                        if j+4 < len(vals): col.setdefault("instructor", j+4)
                        if j+5 < len(vals): col.setdefault("for", j+5)
                        break
                header_idx = i - 1
                break

    if "code" not in col:
        return entries

    data_start = (header_idx or 0) + 1

    current_sem = 0
    for row in rows[data_start:]:
        if not row or all(v is None or _clean(v) == "" for v in row):
            continue

        # Detect semester header rows like "BS MGS Semester 06 (Intake 2023)"
        full_text = " ".join(_clean(v) for v in row)
        sem_match = re.search(r'[Ss]emester\s+(\d+)', full_text)
        if sem_match and not any(
            re.match(r'^[A-Z]{2,3}\d{3}', _clean(row[col.get("code",2)]))
            for _ in [1]
        ):
            current_sem = int(sem_match.group(1))
            continue

        vals = [_clean(v) for v in row]

        def g(key, fallback=""):
            idx = col.get(key)
            if idx is None or idx >= len(vals):
                return fallback
            return vals[idx]

        code = re.sub(r'\s+', '', g("code", "")).upper()
        if not code or not re.match(r'^[A-Z]{1,3}\d', code):
            continue

        title      = g("title")
        ch_raw     = g("ch", "2")
        section    = g("section", "A")
        instructor = g("instructor")
        for_prog   = g("for")
        sem_val    = g("sem") if "sem" in col else current_sem

        # Detect semester from "for" field e.g. "4th"
        if not sem_val:
            sem_m = re.search(r'(\d+)(?:st|nd|rd|th)', full_text)
            if sem_m:
                sem_val = int(sem_m.group(1))

        ch, raw_ch = _parse_ch(ch_raw)
        batch_yr, intake = _sem_to_batch(sem_val or current_sem)

        # Also parse from sheet semester
        if batch_yr == 0:
            sem_sheet = re.search(r'[Ss]emester\s+(\d+)', " ".join(_clean(v) for v in rows[0]))
            if sem_sheet:
                batch_yr, intake = _sem_to_batch(int(sem_sheet.group(1)))

        is_lab = bool(re.search(r'(?:^|[-\s])L$|Lab$|-L$', code, re.I))
        faculty = _infer_faculty(code)

        # num_students from for_prog
        num_students = 30
        sm = re.search(r'=\s*(\d+)', for_prog or "")
        if sm:
            num_students = int(sm.group(1))

        if not section or section in ("None",""):
            section = "A"

        can_name, _ = normalize_teacher_name(instructor) if instructor else ("", "")

        entries.append(CourseEntry(
            course_code  = code,
            title        = title,
            credit_hours = ch,
            instructor   = can_name,
            section      = str(section),
            faculty      = faculty,
            semester     = int(_clean(sem_val)) if str(sem_val).isdigit() else (current_sem or 0),
            batch_year   = batch_yr,
            intake_year  = intake,
            is_lab       = is_lab,
            for_program  = for_prog,
            num_students = num_students,
            raw_ch       = raw_ch,
        ))

    return entries


def parse_csv(filepath: str) -> list[CourseEntry]:
    import csv
    entries = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rl = {k.lower().strip(): v for k, v in row.items()}
            code = re.sub(r'\s+','',_clean(rl.get("code") or rl.get("course code",""))).upper()
            if not code or not re.match(r'^[A-Z]{1,3}\d', code):
                continue
            ch, raw_ch = _parse_ch(rl.get("ch") or rl.get("credit hours","2"))
            section = _clean(rl.get("section","A")) or "A"
            instructor = _clean(rl.get("instructor",""))
            can_name, _ = normalize_teacher_name(instructor) if instructor else ("","")
            sem = rl.get("semester","0")
            batch_yr, intake = _sem_to_batch(sem)
            entries.append(CourseEntry(
                course_code=code, title=_clean(rl.get("title","")),
                credit_hours=ch, instructor=can_name, section=section,
                faculty=_infer_faculty(code), semester=int(_clean(sem) or 0),
                batch_year=batch_yr, intake_year=intake,
                is_lab=bool(re.search(r'(?:^|[-\s])L$|Lab$|-L$',code,re.I)),
                for_program=_clean(rl.get("for","")), raw_ch=raw_ch,
            ))
    return entries


def parse_pdf(filepath: str) -> list[CourseEntry]:
    from pypdf import PdfReader
    reader = PdfReader(filepath)
    text = "\n".join(p.extract_text() or "" for p in reader.pages)
    entries = []
    pattern = re.compile(
        r'([A-Z]{2,3}\s*\d{3}(?:[A-Z]|L)?)\s+'
        r'([A-Za-z][^\d]{5,60?}?)\s+'
        r'(\d+(?:-\d+)*(?:\s*CH)?)\s+'
        r'([A-Z\d]+)\s+([A-Za-z][^\n]{3,50})',
        re.MULTILINE
    )
    for m in pattern.finditer(text):
        code = re.sub(r'\s+','',m.group(1)).upper()
        can_name, _ = normalize_teacher_name(m.group(5).strip())
        ch, raw_ch = _parse_ch(m.group(3).strip())
        entries.append(CourseEntry(
            course_code=code, title=m.group(2).strip(),
            credit_hours=ch, instructor=can_name, section=m.group(4).strip(),
            faculty=_infer_faculty(code),
            is_lab=bool(re.search(r'(?:^|[-\s])L$|Lab$|-L$',code,re.I)),
            raw_ch=raw_ch,
        ))
    return entries


def parse_course_list(filepath: str) -> list[CourseEntry]:
    ext = os.path.splitext(filepath)[1].lower()
    if ext in (".xlsx",".xlsm",".xls",".ods"):
        return parse_excel(filepath)
    elif ext in (".csv",".tsv"):
        return parse_csv(filepath)
    elif ext == ".pdf":
        return parse_pdf(filepath)
    raise ValueError(f"Unsupported: {ext}")


def summarize_courses(entries: list[CourseEntry]) -> dict:
    from collections import Counter
    return {
        "total":       len(entries),
        "by_faculty":  dict(Counter(e.faculty for e in entries)),
        "unique_codes":len(set(e.course_code for e in entries)),
        "labs":        sum(1 for e in entries if e.is_lab),
    }
